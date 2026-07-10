#!/usr/bin/env python3
"""Export Assets/Excel/excel.xlsx into prototype JSON config files."""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("[config] Missing dependency: openpyxl", file=sys.stderr)
    print("[config] Install it with: python -m pip install openpyxl", file=sys.stderr)
    sys.exit(2)


REQUIRED_SHEETS = ("item", "global", "level", "Sheet1")


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def normalize_header(value: Any) -> str:
    value = clean(value)
    if value is None:
        return ""
    return "".join(str(value).split()).lower()


def is_auto_key(value: Any) -> bool:
    value = clean(value)
    if not isinstance(value, str):
        return False
    if not value:
        return False
    return all(char.isascii() and (char.isalnum() or char == "_") for char in value)


def read_header_map(sheet: Any) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        header = sheet.cell(1, column).value
        normalized = normalize_header(header)
        if normalized:
            headers[normalized] = column
    return headers


def get_column_specs(data: Dict[str, Any], sheet: Any) -> List[Dict[str, Any]]:
    specs: List[Dict[str, Any]] = [
        dict(column)
        for column in data.get("columns", [])
        if isinstance(column, dict) and column.get("key")
    ]
    known_keys = {str(column["key"]) for column in specs}
    known_headers = {
        normalize_header(column.get("sourceHeader"))
        for column in specs
        if column.get("sourceHeader")
    }
    for column in range(1, sheet.max_column + 1):
        header = clean(sheet.cell(1, column).value)
        if not is_auto_key(header) or header in known_keys or normalize_header(header) in known_headers:
            continue
        specs.append({"key": header, "sourceHeader": header})
        known_keys.add(str(header))
    return specs


def resolve_column_index(spec: Dict[str, Any], header_map: Dict[str, int]) -> Optional[int]:
    candidates = (spec.get("sourceHeader"), spec.get("key"))
    for candidate in candidates:
        normalized = normalize_header(candidate)
        if normalized and normalized in header_map:
            return header_map[normalized]
    return None


def map_column_specs(data: Dict[str, Any], sheet: Any) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    specs = get_column_specs(data, sheet)
    header_map = read_header_map(sheet)
    key_to_column: Dict[str, int] = {}
    for spec in specs:
        column = resolve_column_index(spec, header_map)
        if column is not None:
            key_to_column[str(spec["key"])] = column
    return specs, key_to_column


def get_row_value(sheet: Any, row_index: int, key_to_column: Dict[str, int], key: str) -> Any:
    column = key_to_column.get(key)
    if column is None:
        return None
    return clean(sheet.cell(row_index, column).value)


def row_has_values(sheet: Any, row_index: int, columns: Iterable[int]) -> bool:
    return any(clean(sheet.cell(row_index, column).value) is not None for column in columns)


def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def save_json(path: Path, data: Dict[str, Any]) -> None:
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def display_path(path: Path, project_root: Path) -> str:
    try:
        return str(path.resolve().relative_to(project_root.resolve())).replace("\\", "/")
    except ValueError:
        return str(path.resolve())


def ensure_required_files(excel_path: Path, output_dir: Path) -> None:
    if not excel_path.exists():
        raise FileNotFoundError(f"Workbook not found: {excel_path}")
    lock_path = excel_path.with_name(f"~${excel_path.name}")
    if lock_path.exists():
        raise PermissionError(
            "Workbook appears to be open in Excel. Save and close it, then run this tool again."
        )
    if not output_dir.exists():
        raise FileNotFoundError(f"Output directory not found: {output_dir}")
    for name in (
        "item_config.json",
        "level_config.json",
        "global_config.json",
        "sheet1_config.json",
    ):
        path = output_dir / name
        if not path.exists():
            raise FileNotFoundError(f"Template JSON not found: {path}")


def ensure_required_sheets(workbook: Any) -> None:
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
    if missing:
        raise ValueError(f"Workbook missing required sheet(s): {', '.join(missing)}")


def export_items(workbook: Any, output_dir: Path, source_workbook: str, now: str) -> int:
    data = load_json(output_dir / "item_config.json")
    items: List[Dict[str, Any]] = []
    sheet = workbook["item"]
    columns, key_to_column = map_column_specs(data, sheet)
    tracked_columns = list(key_to_column.values())

    for row_index in range(2, sheet.max_row + 1):
        if not row_has_values(sheet, row_index, tracked_columns):
            continue
        item_id = get_row_value(sheet, row_index, key_to_column, "itemId")
        if item_id is None:
            continue
        item = {
            str(column["key"]): get_row_value(sheet, row_index, key_to_column, str(column["key"]))
            for column in columns
        }
        items.append(item)

    data["sourceWorkbook"] = source_workbook
    data["generatedAt"] = now
    data["note"] = "Item source weight is stored in kilograms and used directly by gameplay and UI."
    data["sourceSheet"] = "item"
    data["columns"] = columns
    data["items"] = items
    save_json(output_dir / "item_config.json", data)
    return len(items)


def export_levels(
    values_workbook: Any,
    formulas_workbook: Any,
    output_dir: Path,
    source_workbook: str,
    now: str,
) -> int:
    data = load_json(output_dir / "level_config.json")
    levels: List[Dict[str, Any]] = []
    values_sheet = values_workbook["level"]
    formulas_sheet = formulas_workbook["level"]
    columns, key_to_column = map_column_specs(data, values_sheet)
    tracked_columns = list(key_to_column.values())

    for row_index in range(2, values_sheet.max_row + 1):
        if not row_has_values(values_sheet, row_index, tracked_columns):
            continue
        if get_row_value(values_sheet, row_index, key_to_column, "levelId") is None:
            continue

        level = {
            str(column["key"]): get_row_value(values_sheet, row_index, key_to_column, str(column["key"]))
            for column in columns
        }
        speed_column = key_to_column.get("shipSpeedDisplay")
        speed_formula = formulas_sheet.cell(row_index, speed_column).value if speed_column else None
        level["shipSpeedFormula"] = (
            speed_formula
            if isinstance(speed_formula, str) and speed_formula.startswith("=")
            else None
        )
        levels.append(level)

    data["sourceWorkbook"] = source_workbook
    data["generatedAt"] = now
    data["sourceSheet"] = "level"
    data["columns"] = columns
    data["levels"] = levels
    save_json(output_dir / "level_config.json", data)
    return len(levels)


def export_entries(
    workbook: Any,
    output_dir: Path,
    json_name: str,
    sheet_name: str,
    first_row: int,
    source_workbook: str,
    now: str,
) -> int:
    data = load_json(output_dir / json_name)
    sheet = workbook[sheet_name]
    entries = data.get("entries", [])
    key_to_row = {
        str(clean(sheet.cell(row_index, 1).value)): row_index
        for row_index in range(first_row, sheet.max_row + 1)
        if clean(sheet.cell(row_index, 1).value)
    }

    for index, entry in enumerate(entries, start=first_row):
        row_index = key_to_row.get(str(entry.get("key"))) if entry.get("key") else None
        if row_index is None:
            row_index = index
        entry["value"] = clean(sheet.cell(row_index, 2).value)

    known_keys = {entry.get("key") for entry in entries}
    for row_index in range(first_row, sheet.max_row + 1):
        key = clean(sheet.cell(row_index, 1).value)
        if not is_auto_key(key) or key in known_keys:
            continue
        entry: Dict[str, Any] = {"key": key}
        description = clean(sheet.cell(row_index, 3).value)
        if description is not None:
            entry["description"] = description
        entry["value"] = clean(sheet.cell(row_index, 2).value)
        entries.append(entry)
        known_keys.add(key)

    data["sourceWorkbook"] = source_workbook
    data["generatedAt"] = now
    data["sourceSheet"] = sheet_name
    data["parameters"] = {
        entry["key"]: entry.get("value")
        for entry in entries
        if entry.get("key")
    }
    save_json(output_dir / json_name, data)
    return len(entries)


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    project_root = script_dir.parents[1]

    parser = argparse.ArgumentParser(
        description="Export the local Excel workbook into prototype JSON config files."
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=script_dir / "excel.xlsx",
        help="Path to the source workbook. Defaults to Assets/Excel/excel.xlsx.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=project_root / "Assets" / "Resources_Prototype" / "TestData",
        help="Directory that contains the prototype JSON config files.",
    )
    args = parser.parse_args()

    excel_path = args.excel.resolve()
    output_dir = args.output.resolve()
    ensure_required_files(excel_path, output_dir)

    values_workbook = load_workbook(excel_path, data_only=True, read_only=True)
    formulas_workbook = load_workbook(excel_path, data_only=False, read_only=True)
    ensure_required_sheets(values_workbook)
    ensure_required_sheets(formulas_workbook)

    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    source_workbook = display_path(excel_path, project_root)

    item_count = export_items(values_workbook, output_dir, source_workbook, now)
    level_count = export_levels(values_workbook, formulas_workbook, output_dir, source_workbook, now)
    global_count = export_entries(
        values_workbook,
        output_dir,
        "global_config.json",
        "global",
        2,
        source_workbook,
        now,
    )
    sheet1_count = export_entries(
        values_workbook,
        output_dir,
        "sheet1_config.json",
        "Sheet1",
        1,
        source_workbook,
        now,
    )

    print(f"[config] Source: {source_workbook}")
    print(f"[config] Output: {display_path(output_dir, project_root)}")
    print(f"[config] Items: {item_count}")
    print(f"[config] Levels: {level_count}")
    print(f"[config] Global entries: {global_count}")
    print(f"[config] Sheet1 entries: {sheet1_count}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"[config] Export failed: {exc}", file=sys.stderr)
        raise SystemExit(1)
