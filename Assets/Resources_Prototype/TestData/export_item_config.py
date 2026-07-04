import argparse
import json
import random
from pathlib import Path

import openpyxl


DEFAULT_EXCEL_PATH = Path(r"C:\Users\windchaser\Downloads\配置表.xlsx")
DEFAULT_OUTPUT_PATH = Path("Assets/Resources_Prototype/TestData/item_config.json")
PIECE_PREFAB_FOLDER = Path("Assets/Prefabs/Pieces")
ITEM_SHEET_NAME = "item"
GLOBAL_SHEET_NAME = "global"

GLOBAL_FIELDS = [
    ("sustainDurationSeconds", "坚持时间，单位秒。模拟持续达到该时间且船未进入危险区则成功。", 20.0, "float", "坚持时间(秒）"),
    ("dropSpeedLimitMetersPerSecond", "掉落速度软上限，单位 m/s。超过后会施加向上阻力。", 4.0, "float", "掉落速度上限（m/s)"),
    ("waterSurfaceTensionCoefficient", "水面张力系数。控制穿过水面时的竖向水面力。", 1.0, "float", "水面张力系数"),
    ("forceCoefficient", "受力系数。控制绳子拉船/反拉相关的全局力倍率。", 1.0, "float", "受力系数"),
    ("itemMassScale", "物品质量系数。物品重量转换为 Unity Rigidbody2D.mass 的全局倍率。", 1.0, "float", "物品质量系数"),
    ("weightForceScale", "重量下压力系数。物品重量转换为向下力的倍率。", 0.008, "float", "重量下压力系数"),
    ("waterDragForceScale", "水阻系数。物品水阻参与速度衰减的全局倍率。", 0.32, "float", "水阻系数"),
    ("thresholdCoefficient", "阈值系数。防御值转换为连接受力阈值的倍率。", 3.0, "float", "阈值系数"),
    ("damageCoefficient", "伤害系数。超过连接阈值后的扣血速度倍率。", 1.0, "float", "伤害系数"),
    ("healthCoefficient", "血量系数。物品健康值转换为连接最大血量的倍率。", 0.3, "float", "血量系数"),
    ("jointFrequencyPerDefense", "防御转关节频率系数。防御越高，Unity FixedJoint2D 越硬。", 0.025, "float", "防御转关节频率系数"),
    ("seabedFriction", "海底 Unity 物理材质摩擦系数。", 0.70, "float", "海底摩擦系数"),
    ("seabedBounciness", "海底 Unity 物理材质弹性系数。", 0.02, "float", "海底弹性系数"),
    ("showJointHealthDebug", "是否显示连接血量调试信息。1 显示，0 隐藏。", 1, "int", "显示连接血量调试"),
]


def as_int(value, default=0):
    if value is None or value == "":
        return default
    return int(round(float(value)))


def as_float(value, default=0.0):
    if value is None or value == "":
        return default
    return float(value)


def as_text(value):
    if value is None:
        return ""
    return str(value).strip()


def make_item_id(raw_id):
    return f"item_{as_int(raw_id):03d}"


def load_items(excel_path):
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    if ITEM_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Missing sheet: {ITEM_SHEET_NAME}")

    sheet = workbook[ITEM_SHEET_NAME]
    items = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = list(row)
        item_id = values[0]
        if item_id is None:
            continue

        item = {
            "itemId": as_int(values[0]),
            "id": make_item_id(values[0]),
            "displayName": as_text(values[1]),
            "artResource": as_text(values[2]),
            "weightGrams": as_float(values[3]),
            "defense": as_float(values[4]),
            "health": as_float(values[5]),
            "friction": as_float(values[10] if len(values) > 10 else None, 0.5),
            "sizeCoefficient": as_float(values[6], 1.0),
            "guaranteedSpawnCount": as_int(values[7]),
            "randomWeight": as_int(values[8]),
            "unlockLevel": as_int(values[9], 1),
        }
        items.append(item)

    return items


def load_globals(excel_path):
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    if GLOBAL_SHEET_NAME not in workbook.sheetnames:
        return default_globals()

    values = {}
    sheet = workbook[GLOBAL_SHEET_NAME]
    for row in sheet.iter_rows(min_row=1, values_only=True):
        key = as_text(row[0])
        if not key:
            continue
        if key == "参数名":
            continue
        values[key] = row[2] if len(row) >= 3 else row[1]

    globals_config = {}
    for key, _, default, value_type, legacy_key in GLOBAL_FIELDS:
        raw_value = values.get(key, values.get(legacy_key))
        globals_config[key] = as_int(raw_value, default) if value_type == "int" else as_float(raw_value, default)

    return globals_config


def default_globals():
    return {key: default for key, _, default, _, _ in GLOBAL_FIELDS}


def assign_prefab_resources(items):
    prefab_paths = [path.as_posix() for path in PIECE_PREFAB_FOLDER.glob("*.prefab")]
    if not prefab_paths:
        return

    random.shuffle(prefab_paths)
    for index, item in enumerate(items):
        if item.get("artResource"):
            continue
        item["artResource"] = prefab_paths[index % len(prefab_paths)]


def main():
    parser = argparse.ArgumentParser(description="Export Broken Anchor item config JSON from 配置表.xlsx.")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL_PATH), help="Path to 配置表.xlsx.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT_PATH), help="Output JSON path.")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    output_path = Path(args.output)
    globals_config = load_globals(excel_path)
    items = load_items(excel_path)
    assign_prefab_resources(items)

    payload = {"globals": globals_config, "items": items}
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Exported {len(items)} items to {output_path}")


if __name__ == "__main__":
    main()
